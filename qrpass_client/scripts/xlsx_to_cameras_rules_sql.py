from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl


def esc(value: str) -> str:
    return value.replace("'", "''")


def normalize_marker(raw: object) -> str:
    if raw is None:
        return ""
    return str(raw).strip().upper()


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert XLSX to SQL for Cameras + Rules.")
    parser.add_argument("--xlsx", required=True, help="Path to xlsx file")
    parser.add_argument("--out", required=True, help="Output SQL path")
    parser.add_argument("--sheet", default=None, help="Sheet name (optional)")
    parser.add_argument("--name-col", type=int, default=3, help="1-based camera name column")
    parser.add_argument("--rule-col", type=int, default=5, help="1-based rules marker column")
    parser.add_argument("--url-col", type=int, default=9, help="1-based rtsp url column")
    parser.add_argument("--start-row", type=int, default=2, help="First data row")
    parser.add_argument("--red-color-id", type=int, default=2, help="Color id for RED/NotRED rules")
    args = parser.parse_args()

    xlsx = Path(args.xlsx).expanduser().resolve()
    out = Path(args.out).expanduser().resolve()

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    rows: list[tuple[str, str, str]] = []
    for row_idx in range(args.start_row, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=args.name_col).value
        marker = ws.cell(row=row_idx, column=args.rule_col).value
        url = ws.cell(row=row_idx, column=args.url_col).value
        if not name or not url:
            continue

        name_s = str(name).strip()
        url_s = str(url).strip()
        if not name_s or not url_s.lower().startswith("rtsp://"):
            continue

        rows.append((name_s, url_s, normalize_marker(marker)))

    lines = [
        "-- Auto-generated from XLSX",
        "-- Rules mapping:",
        "--   RED -> color_id=<red-color-id>, access_granted=1",
        "--   NOTRED/NOTERED -> color_id=<red-color-id>, access_granted=0",
        "--   BLUE/other/empty -> no rows in Rules",
        "BEGIN TRANSACTION;",
        "",
    ]

    for name, url, marker in rows:
        eu = esc(url)
        en = esc(name)
        lines.append(
            "INSERT INTO Cameras (url, name) "
            f"SELECT '{eu}', '{en}' "
            f"WHERE NOT EXISTS (SELECT 1 FROM Cameras WHERE url = '{eu}');"
        )
        lines.append(f"UPDATE Cameras SET name = '{en}' WHERE url = '{eu}';")
        lines.append(
            f"DELETE FROM Rules WHERE camera_id = (SELECT id FROM Cameras WHERE url = '{eu}' LIMIT 1);"
        )
        if marker == "RED":
            lines.append(
                "INSERT INTO Rules (camera_id, color_id, access_granted) "
                f"SELECT id, {args.red_color_id}, 1 FROM Cameras WHERE url = '{eu}' LIMIT 1;"
            )
        elif marker in ("NOTRED", "NOTERED"):
            lines.append(
                "INSERT INTO Rules (camera_id, color_id, access_granted) "
                f"SELECT id, {args.red_color_id}, 0 FROM Cameras WHERE url = '{eu}' LIMIT 1;"
            )
        lines.append("")

    lines.append("COMMIT;")
    lines.append("")

    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text("\n".join(lines), encoding="utf-8")

    print(f"Generated: {out}")
    print(f"Rows: {len(rows)}")


if __name__ == "__main__":
    main()
