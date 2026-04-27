from __future__ import annotations

import argparse
import sqlite3
import time
from pathlib import Path

import openpyxl


def _norm_header(v: object) -> str:
    return "".join(ch for ch in str(v or "").strip().lower() if ch.isalnum())


def _truthy(v: object) -> bool:
    s = str(v or "").strip().lower()
    return s in {"1", "true", "yes", "y", "on", "да", "x"}


def main() -> None:
    p = argparse.ArgumentParser(description="Импорт списка камер подсчета свиней из XLSX в users.db")
    p.add_argument("--xlsx", required=True, help="Путь к xlsx")
    p.add_argument("--db", default="users.db", help="Путь к users.db")
    p.add_argument("--sheet", default=None, help="Имя листа (опционально)")
    p.add_argument("--replace", action="store_true", help="Полностью заменить список в pig_count_cameras")
    args = p.parse_args()

    xlsx = Path(args.xlsx).expanduser().resolve()
    db_path = Path(args.db).expanduser().resolve()
    if not xlsx.is_file():
        raise SystemExit(f"XLSX не найден: {xlsx}")
    if not db_path.is_file():
        raise SystemExit(f"DB не найден: {db_path}")

    wb = openpyxl.load_workbook(str(xlsx), data_only=True, read_only=True)
    ws = wb[args.sheet] if (args.sheet and args.sheet in wb.sheetnames) else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        raise SystemExit("Пустой лист.")
    idx = {_norm_header(v): i for i, v in enumerate(header)}

    cam_idx = None
    for k in ("cameraname", "camera", "name", "камера", "имякамеры"):
        if k in idx:
            cam_idx = idx[k]
            break
    if cam_idx is None:
        raise SystemExit("Не найдена колонка camera_name/camera/name.")

    enabled_idx = None
    for k in ("enabled", "isenabled", "pigcountenabled", "countenabled", "use", "включено"):
        if k in idx:
            enabled_idx = idx[k]
            break

    selected: list[str] = []
    for r in rows:
        if r is None:
            continue
        cam = str(r[cam_idx] if cam_idx < len(r) else "").strip()
        if not cam:
            continue
        if enabled_idx is not None:
            v = r[enabled_idx] if enabled_idx < len(r) else ""
            if not _truthy(v):
                continue
        selected.append(cam)

    conn = sqlite3.connect(str(db_path))
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS pig_count_cameras (
                camera_name TEXT PRIMARY KEY,
                enabled INTEGER NOT NULL DEFAULT 1,
                updated_at REAL NOT NULL DEFAULT (strftime('%s','now'))
            )
            """
        )
        if args.replace:
            conn.execute("DELETE FROM pig_count_cameras")
        now = time.time()
        for cam in sorted(set(selected)):
            conn.execute(
                """
                INSERT INTO pig_count_cameras(camera_name, enabled, updated_at)
                VALUES(?, 1, ?)
                ON CONFLICT(camera_name) DO UPDATE SET enabled=1, updated_at=excluded.updated_at
                """,
                (cam, now),
            )
        conn.commit()
    finally:
        conn.close()

    print(f"Импортировано камер: {len(set(selected))}")
    print(f"DB: {db_path}")


if __name__ == "__main__":
    main()
