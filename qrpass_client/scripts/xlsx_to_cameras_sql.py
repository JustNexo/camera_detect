from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl


def esc(value: str) -> str:
    return value.replace("'", "''")


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert camera XLSX to SQL upserts for Cameras table.")
    parser.add_argument("--xlsx", required=True, help="Path to xlsx file")
    parser.add_argument("--out", default="sql/update_cameras_from_xlsx.sql", help="Output SQL path")
    parser.add_argument("--sheet", default=None, help="Sheet name (optional)")
    parser.add_argument("--name-col", type=int, default=8, help="1-based camera name column")
    parser.add_argument("--url-col", type=int, default=9, help="1-based rtsp url column")
    parser.add_argument("--start-row", type=int, default=2, help="First data row")
    args = parser.parse_args()

    xlsx = Path(args.xlsx).expanduser().resolve()
    out = Path(args.out).expanduser().resolve()

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    rows: list[tuple[str, str]] = []
    for row_idx in range(args.start_row, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=args.name_col).value
        url = ws.cell(row=row_idx, column=args.url_col).value
        if not name or not url:
            continue
        name_s = str(name).strip()
        url_s = str(url).strip()
        if not name_s or not url_s.lower().startswith("rtsp://"):
            continue
        rows.append((name_s, url_s))

    lines = [
        "-- Auto-generated from XLSX",
        "BEGIN TRANSACTION;",
    ]
    for name, url in rows:
        lines.append(
            "INSERT INTO Cameras (url, name) "
            f"SELECT '{esc(url)}', '{esc(name)}' "
            f"WHERE NOT EXISTS (SELECT 1 FROM Cameras WHERE url = '{esc(url)}');"
        )
        lines.append(
            f"UPDATE Cameras SET name = '{esc(name)}' WHERE url = '{esc(url)}';"
        )
    lines.append("COMMIT;")
    lines.append("")

    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text("\n".join(lines), encoding="utf-8")
    print(f"Generated: {out}")
    print(f"Rows: {len(rows)}")


if __name__ == "__main__":
    main()
